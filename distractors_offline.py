#!/usr/bin/env python3
"""Add two more wrong answers (incorrect_answer_4/5) to every question in opendb.xlsx.

Fully offline and free — no API, no model. For each question it draws two extra
distractors from OTHER questions in the SAME category whose answer has the same
format (numbers paired with numbers, text with text of similar length), skipping
anything equal to the correct answer or the existing wrong answers. Falls back to
looser matching only when a category doesn't have enough candidates.
"""

import os
import random
import re

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "opendb.xlsx")

# Original column positions (1-based). Scoring may have added columns after these.
COL_CAT, COL_Q, COL_CORRECT, COL_W1, COL_W2, COL_W3 = 4, 5, 6, 7, 8, 9

_NUM_RE = re.compile(r"^[\d.,]+$")


def norm(s):
    return (s or "").strip().lower()


def shape(answer):
    """Coarse format bucket so distractors look like the real answers."""
    a = (answer or "").strip()
    if not a:
        return "text"
    if _NUM_RE.match(a) and any(c.isdigit() for c in a):
        return "num"
    # length band for text so a one-word answer isn't paired with a long phrase
    n = len(a)
    if n <= 12:
        return "text-short"
    if n <= 30:
        return "text-mid"
    return "text-long"


def main():
    import shutil
    from openpyxl import load_workbook

    shutil.copy(XLSX, os.path.join(HERE, "opendb_predistractor_backup.xlsx"))
    wb = load_workbook(XLSX)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # --- Build answer pools per category and per (category, shape).
    by_cat = {}          # category -> set of answer strings
    by_cat_shape = {}    # (category, shape) -> list of answer strings
    seen_global = {}     # shape -> list (fallback pool)
    for r in rows:
        cat = (r[COL_CAT - 1] or "").strip()
        answers = [r[COL_CORRECT - 1]] + [r[c - 1] for c in (COL_W1, COL_W2, COL_W3)]
        for a in answers:
            a = (a or "").strip()
            if not a:
                continue
            by_cat.setdefault(cat, set()).add(a)
            sh = shape(a)
            by_cat_shape.setdefault((cat, sh), []).append(a)
            seen_global.setdefault(sh, []).append(a)

    def pick_two(cat, existing_norm, target_shape, seed):
        """Two distinct distractors, preferring same category + same shape."""
        rng = random.Random(seed)
        chosen = []
        chosen_norm = set(existing_norm)

        def draw_from(pool):
            cands = [a for a in pool if norm(a) not in chosen_norm]
            rng.shuffle(cands)
            for a in cands:
                if len(chosen) >= 2:
                    break
                chosen.append(a)
                chosen_norm.add(norm(a))

        # 1) same category, same shape
        draw_from(by_cat_shape.get((cat, target_shape), []))
        # 2) same category, any shape
        if len(chosen) < 2:
            draw_from(list(by_cat.get(cat, set())))
        # 3) any category, same shape
        if len(chosen) < 2:
            draw_from(seen_global.get(target_shape, []))
        # 4) any category, any shape (last resort)
        if len(chosen) < 2:
            draw_from([a for pool in seen_global.values() for a in pool])
        return chosen[0] if chosen else "", chosen[1] if len(chosen) > 1 else ""

    # --- Write two new columns after whatever already exists.
    base = ws.max_column
    ws.cell(row=1, column=base + 1, value="incorrect_answer_4")
    ws.cell(row=1, column=base + 2, value="incorrect_answer_5")

    filled = 0
    for i, r in enumerate(rows):
        cat = (r[COL_CAT - 1] or "").strip()
        correct = (r[COL_CORRECT - 1] or "").strip()
        existing = [correct] + [(r[c - 1] or "").strip() for c in (COL_W1, COL_W2, COL_W3)]
        existing_norm = {norm(a) for a in existing if a}
        target_shape = shape(correct)
        # deterministic per-question seed so re-runs are stable
        seed = hash((cat, (r[COL_Q - 1] or "")))
        d4, d5 = pick_two(cat, existing_norm, target_shape, seed)
        ws.cell(row=i + 2, column=base + 1, value=d4)
        ws.cell(row=i + 2, column=base + 2, value=d5)
        if d4 and d5:
            filled += 1

    wb.save(XLSX)
    print(f"Done. Added 2 distractors to {filled}/{len(rows)} questions in {XLSX}")


if __name__ == "__main__":
    main()
