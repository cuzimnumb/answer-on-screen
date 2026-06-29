#!/usr/bin/env python3
"""Fetch trivia questions from the Open Trivia Database API and append them to opendb.xlsx.

Hits https://opentdb.com/api.php?amount=50 once every 1.5 seconds for 150 rounds,
decoding each batch of questions and writing one question per row.
"""

import html
import json
import os
import ssl
import time
import urllib.error
import urllib.request

API_URL = "https://opentdb.com/api.php?amount=50"
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "opendb.xlsx")
ROUNDS = 150
INTERVAL_SECONDS = 5.0

# Skip these when writing rows.
EXCLUDE_CATEGORIES = {"Entertainment: Video Games"}
EXCLUDE_TYPES = {"boolean"}  # boolean == True/False questions


def question_key(question):
    """Normalized question text used to detect duplicates."""
    return (question or "").strip().lower()


def is_excluded(q):
    return (
        q.get("category") in EXCLUDE_CATEGORIES
        or q.get("type") in EXCLUDE_TYPES
    )

HEADERS = [
    "round",
    "type",
    "difficulty",
    "category",
    "question",
    "correct_answer",
    "incorrect_answer_1",
    "incorrect_answer_2",
    "incorrect_answer_3",
]


def unescape(value):
    """API returns HTML-encoded text (e.g. &quot;, &#039;); decode to plain text."""
    return html.unescape(value) if isinstance(value, str) else value


# Verified TLS by default. Some Python installs on macOS ship without a CA bundle,
# in which case verification fails; we then fall back to an unverified context, which
# is acceptable for this public, read-only trivia API. _UNVERIFIED is set lazily on the
# first SSL failure so the verified path is always tried first.
_UNVERIFIED = None


def _unverified_context():
    global _UNVERIFIED
    if _UNVERIFIED is None:
        print("Warning: CA verification unavailable; using an unverified TLS context.")
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        _UNVERIFIED = ctx
    return _UNVERIFIED


def fetch_round():
    """Fetch one batch of 50 questions; returns the list of result dicts."""
    req = urllib.request.Request(API_URL, headers={"User-Agent": "opendb-fetcher/1.0"})
    try:
        resp = urllib.request.urlopen(req, timeout=30)
    except urllib.error.URLError as exc:
        # urlopen wraps cert failures in URLError(reason=SSLError); retry unverified.
        if not isinstance(exc.reason, ssl.SSLError):
            raise
        resp = urllib.request.urlopen(req, timeout=30, context=_unverified_context())
    with resp:
        data = json.loads(resp.read().decode("utf-8"))
    # response_code 0 == success per the OpenTDB API spec
    if data.get("response_code") != 0:
        return []
    return data.get("results", [])


def main():
    from openpyxl import load_workbook, Workbook

    # Track questions already written so we never add duplicates (across the whole file).
    seen = set()

    # Resume: append to an existing opendb.xlsx if present; otherwise start fresh.
    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        existing = ws.max_row - 1  # minus header
        for row in ws.iter_rows(min_row=2, values_only=True):
            seen.add(question_key(row[4]))  # column 5 == question text
        # Continue round numbering past the last batch already written.
        start_round = (existing // 50) + 1
        print(f"Resuming {OUTPUT_FILE}: {existing} rows present, continuing at round {start_round}.")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "trivia"
        ws.append(HEADERS)
        start_round = 1

    end_round = start_round + ROUNDS - 1
    written = 0
    for round_num in range(start_round, end_round + 1):
        try:
            results = fetch_round()
        except Exception as exc:  # network hiccup / rate limit — log and keep going
            print(f"[round {round_num}/{end_round}] error: {exc}")
            results = []

        kept = 0
        for q in results:
            if is_excluded(q):
                continue
            key = question_key(unescape(q.get("question", "")))
            if key in seen:
                continue  # duplicate of something already in the file
            seen.add(key)
            incorrect = [unescape(a) for a in q.get("incorrect_answers", [])]
            # pad to 3 columns so rows stay aligned
            incorrect += [""] * (3 - len(incorrect))
            ws.append([
                round_num,
                unescape(q.get("type", "")),
                unescape(q.get("difficulty", "")),
                unescape(q.get("category", "")),
                unescape(q.get("question", "")),
                unescape(q.get("correct_answer", "")),
                incorrect[0],
                incorrect[1],
                incorrect[2],
            ])
            kept += 1

        written += kept
        skipped = len(results) - kept
        print(f"[round {round_num}/{end_round}] +{kept} kept, {skipped} filtered (added {written})")

        # Save after every round so progress survives an interruption.
        wb.save(OUTPUT_FILE)

        if round_num < end_round:
            time.sleep(INTERVAL_SECONDS)

    print(f"Done. Added {written} questions to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
