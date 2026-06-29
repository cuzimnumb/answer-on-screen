#!/usr/bin/env python3
"""Add popularity_score, numeric_difficulty and outlier columns to opendb.xlsx.

popularity_score for a question = 50% * Wikipedia popularity of the correct answer
                                + 50% * mean Wikipedia popularity of the wrong answers.
"Popularity" = total English-Wikipedia pageviews over the last 12 months for the
article that best matches the answer text (resolved via the search API).

Lower popularity_score => harder question. Within each existing easy/medium/hard band
the questions are ranked by score and split into numeric difficulty levels:
    easy  -> 1,2,3   (lowest score in band = 3, the hardest easy)
    medium-> 4,5     (lowest score = 5)
    hard  -> 6,7     (lowest score = 7)

Special cases:
  * Questions whose answers are ALL purely numeric get no popularity lookup and are
    assigned the hardest level of their band (easy=3, medium=5, hard=7).
  * If the correct answer has no Wikipedia page, the score falls back to the mean of
    whatever wrong answers could be resolved.

Outliers (score very different from peers in the same category+difficulty group) are
flagged TRUE in the `outlier` column using the IQR rule.

Wikipedia lookups are cached in wiki_cache.json so re-runs are cheap and resumable.
"""

import json
import os
import re
import ssl
import statistics
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed

WORKERS = 16  # concurrent Wikipedia lookups (latency-bound, so concurrency helps a lot)

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "opendb.xlsx")
CACHE_FILE = os.path.join(HERE, "wiki_cache.json")

# Last 12 complete months (today is mid-2026): 2025-06 .. 2026-05 inclusive.
PV_START = "2025060100"
PV_END = "2026053100"

# Wikimedia throttles (HTTP 429) requests without a descriptive, contactable UA.
USER_AGENT = (
    "opendb-trivia-difficulty/1.0 "
    "(https://github.com/local/opendb; sk8my8yhkg@privaterelay.appleid.com) python-urllib"
)

# Column layout already in the sheet (1-based): question text is col 5, etc.
COL_TYPE, COL_DIFF, COL_CAT, COL_Q = 2, 3, 4, 5
COL_CORRECT, COL_W1, COL_W2, COL_W3 = 6, 7, 8, 9

BAND_LEVELS = {  # ordered hardest -> easiest; index 0 == lowest score == hardest
    "easy": [3, 2, 1],
    "medium": [5, 4],
    "hard": [7, 6],
}

_ctx = ssl.create_default_context()
_ctx.check_hostname = False
_ctx.verify_mode = ssl.CERT_NONE

_NUMERIC_RE = re.compile(r"^[\d.,\s]+$")


def is_numeric(answer):
    """True if the answer is a pure number (digits/commas/dots only), e.g. '1997', '4'."""
    a = (answer or "").strip()
    return bool(a) and bool(_NUMERIC_RE.match(a)) and any(c.isdigit() for c in a)


def _get_json(url):
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    for attempt in range(4):
        try:
            with urllib.request.urlopen(req, timeout=30, context=_ctx) as r:
                return json.loads(r.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            if e.code == 404:
                return None  # no such article / no stats
            if e.code == 429:  # rate limited — back off and retry
                time.sleep(2 * (attempt + 1))
                continue
            raise
        except urllib.error.URLError:
            time.sleep(1 + attempt)
    return None


def resolve_title(term):
    url = "https://en.wikipedia.org/w/api.php?" + urllib.parse.urlencode({
        "action": "query", "list": "search", "srsearch": term,
        "srlimit": 1, "format": "json",
    })
    data = _get_json(url)
    if not data:
        return None
    hits = data.get("query", {}).get("search", [])
    return hits[0]["title"] if hits else None


def page_views(title):
    t = urllib.parse.quote(title.replace(" ", "_"), safe="")
    url = (
        "https://wikimedia.org/api/rest_v1/metrics/pageviews/per-article/"
        f"en.wikipedia/all-access/all-agents/{t}/monthly/{PV_START}/{PV_END}"
    )
    data = _get_json(url)
    if not data:
        return None
    return sum(item.get("views", 0) for item in data.get("items", []))


def load_cache():
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(cache):
    tmp = CACHE_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(cache, f)
    os.replace(tmp, CACHE_FILE)


def fetch_answer(answer):
    """Resolve + fetch pageviews for one answer; returns (answer, title, views)."""
    title = resolve_title(answer)
    views = page_views(title) if title else None
    return answer, title, views


def popularity(answer, cache):
    """Cached pageviews for an answer string; None if no page could be resolved."""
    key = answer.strip()
    if key in cache:
        return cache[key]["views"]
    _, title, views = fetch_answer(key)
    cache[key] = {"title": title, "views": views}
    return views


def prefetch(answers, cache):
    """Populate the cache for all answer strings concurrently."""
    todo = sorted({a.strip() for a in answers if a.strip()} - cache.keys())
    if not todo:
        return
    print(f"Prefetching {len(todo)} uncached answers with {WORKERS} workers...")
    done = 0
    with ThreadPoolExecutor(max_workers=WORKERS) as pool:
        futures = {pool.submit(fetch_answer, a): a for a in todo}
        for fut in as_completed(futures):
            answer = futures[fut]
            try:
                _, title, views = fut.result()
            except Exception:
                title, views = None, None
            cache[answer] = {"title": title, "views": views}
            done += 1
            if done % 200 == 0:
                save_cache(cache)
                print(f"  {done}/{len(todo)} answers fetched...")
    save_cache(cache)
    print(f"  {done}/{len(todo)} answers fetched. Done prefetching.")


def main():
    import shutil
    from openpyxl import load_workbook

    shutil.copy(XLSX, os.path.join(HERE, "opendb_prescore_backup.xlsx"))
    wb = load_workbook(XLSX)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    cache = load_cache()

    # --- Pass 1: resolve popularity for every answer of every (non all-numeric) question.
    questions = []  # dicts holding computed fields, in sheet order
    pending = 0
    for i, r in enumerate(rows):
        correct = (r[COL_CORRECT - 1] or "").strip()
        wrongs = [(r[c - 1] or "").strip() for c in (COL_W1, COL_W2, COL_W3)]
        wrongs = [w for w in wrongs if w]
        answers = [correct] + wrongs
        all_numeric = answers and all(is_numeric(a) for a in answers)

        q = {
            "band": (r[COL_DIFF - 1] or "").strip().lower(),
            "category": (r[COL_CAT - 1] or "").strip(),
            "correct": correct,
            "wrongs": wrongs,
            "all_numeric": all_numeric,
            "score": None,
        }
        questions.append(q)
        if not all_numeric:
            pending += 1

    print(f"{len(questions)} questions; {pending} need Wikipedia lookups.")

    # Prefetch every answer of the non all-numeric questions, concurrently.
    needed = []
    for q in questions:
        if q["all_numeric"]:
            continue
        needed.append(q["correct"])
        needed.extend(q["wrongs"])
    prefetch(needed, cache)

    done = 0
    for idx, q in enumerate(questions):
        if q["all_numeric"]:
            continue
        pop_correct = popularity(q["correct"], cache)
        wrong_pops = [p for p in (popularity(w, cache) for w in q["wrongs"]) if p is not None]
        wrong_mean = statistics.mean(wrong_pops) if wrong_pops else None

        if pop_correct is None:
            # Fallback: mean of resolvable wrong answers (or 0 if none resolvable).
            q["score"] = wrong_mean if wrong_mean is not None else 0.0
        elif wrong_mean is not None:
            q["score"] = 0.5 * pop_correct + 0.5 * wrong_mean
        else:
            q["score"] = float(pop_correct)  # no wrongs resolvable -> just the correct

        done += 1
        if done % 50 == 0:
            save_cache(cache)
            print(f"  scored {done}/{pending}...")
    save_cache(cache)

    # --- Pass 2: numeric difficulty per band (rank by score; lowest score = hardest).
    for band, levels in BAND_LEVELS.items():
        scored = [q for q in questions if q["band"] == band and not q["all_numeric"]]
        scored.sort(key=lambda q: q["score"])
        n = len(scored)
        for pos, q in enumerate(scored):
            grp = min(pos * len(levels) // n, len(levels) - 1) if n else 0
            q["numeric_difficulty"] = levels[grp]
        # All-numeric questions -> hardest level of their band.
        for q in questions:
            if q["band"] == band and q["all_numeric"]:
                q["numeric_difficulty"] = levels[0]

    # Any band not in BAND_LEVELS (shouldn't happen) -> leave blank
    for q in questions:
        q.setdefault("numeric_difficulty", None)

    # --- Pass 3: outliers within each (category, difficulty) group via IQR.
    groups = {}
    for q in questions:
        if q["all_numeric"] or q["score"] is None:
            q["outlier"] = False
            continue
        groups.setdefault((q["category"], q["band"]), []).append(q)

    for members in groups.values():
        scores = sorted(m["score"] for m in members)
        for m in members:
            m["outlier"] = False
        if len(scores) < 5:
            continue  # too few to judge an outlier meaningfully
        q1, q3 = _quantile(scores, 0.25), _quantile(scores, 0.75)
        iqr = q3 - q1
        lo, hi = q1 - 1.5 * iqr, q3 + 1.5 * iqr
        for m in members:
            if m["score"] < lo or m["score"] > hi:
                m["outlier"] = True

    # --- Write new columns.
    headers = [c.value for c in ws[1]]
    base = len(headers)
    ws.cell(row=1, column=base + 1, value="popularity_score")
    ws.cell(row=1, column=base + 2, value="numeric_difficulty")
    ws.cell(row=1, column=base + 3, value="outlier")

    n_out = 0
    for i, q in enumerate(questions):
        row = i + 2
        score = None if q["score"] is None else round(q["score"])
        ws.cell(row=row, column=base + 1, value=score)
        ws.cell(row=row, column=base + 2, value=q.get("numeric_difficulty"))
        ws.cell(row=row, column=base + 3, value="TRUE" if q.get("outlier") else "")
        if q.get("outlier"):
            n_out += 1

    wb.save(XLSX)
    print(f"Done. Wrote columns for {len(questions)} questions; {n_out} flagged as outliers.")


def _quantile(sorted_vals, q):
    """Linear-interpolation quantile (same method as numpy default)."""
    if not sorted_vals:
        return 0.0
    pos = (len(sorted_vals) - 1) * q
    lo = int(pos)
    hi = min(lo + 1, len(sorted_vals) - 1)
    frac = pos - lo
    return sorted_vals[lo] * (1 - frac) + sorted_vals[hi] * frac


if __name__ == "__main__":
    main()
